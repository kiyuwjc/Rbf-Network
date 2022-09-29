import pandas
import numpy as np
import xlrd2
import openpyxl as op
data = xlrd2.open_workbook('数据.xls')
table = data.sheet_by_name('Sheet1')
row_num = table.nrows - 1
col_num = table.ncols - 1
x = np.zeros((row_num, col_num))



constant = {
    "atm": 101, "g": 10, "s_vap": 3.17
}
# # print(x[0, 0])
# print(table.cell(120, 4).value)
for i in range(1, row_num + 1):
    for j in range(1, col_num + 1):
        # print(i, j)
        x[i - 1, j - 1] = table.cell(i, j).value
def q_bi(vpa):
    q = []
    for i in vpa:
        q_ = 0.662 * (i / 1000) / (constant["atm"] - 0.378 * (i / 1000))
        q.append(q_)
        pass
    return q
    pass

q = q_bi(x[:, 0])

def re_(vap):
    re = []
    for i in vap:
        r_ = i / (1000 * constant["s_vap"])
        re.append(r_)
    return re

r = re_(x[:, 0])

def v_vap(v, q):
    v_ = []
    for i in range(row_num):
        result = (abs(v[i]) * q[i]) / constant["g"]
        v_.append(result)
        pass
    return v_

v = v_vap(x[:, 2], q)


def pre(pr):
    p = []
    for i in pr:
        p.append(i)
        pass
    return p

p = pre(x[:, 3])

wb = op.Workbook()
ws = wb['Sheet']
num1 = 1

for i in q:
    ws.cell(row=num1, column=1).value = i
    num1 = num1 + 1
    pass

num2 = 1

for i in r:
    ws.cell(row=num2, column=2).value = i
    num2 = num2 + 1
    pass

num3 = 1

for i in v:
    ws.cell(row=num3, column=3).value = i
    num3 = num3 + 1
    pass

num4 = 1
for i in p:
    ws.cell(row=num4, column=4).value = i
    num4 = num4 + 1
    pass
wb.save('123.xls')