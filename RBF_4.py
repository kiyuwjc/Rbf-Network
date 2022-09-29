import numpy as np
from matplotlib import pyplot as plt
import matplotlib as mpl
from scipy.linalg import norm, pinv
import xlrd2
import openpyxl as op
from pandas import DataFrame as df
# np.random.seed(0)
np.set_printoptions(suppress=True)
class RBF:
    def __init__(self, input_dim, num_centers, out_dim):
        self.input_dim = input_dim
        self.num_centers = num_centers
        self.out_dim = out_dim
        self.beta = 85000
        self.centers = [np.random.uniform(-1, 1, input_dim) for i in range(num_centers)]
        self.w = np.random.random((self.num_centers, self.out_dim))

    def basic_func(self, c, d):
        return np.exp(-self.beta * norm(c - d) ** 2)

    def act(self, x):
        g = np.zeros((x.shape[0], self.num_centers), dtype=np.float16)
        for ci, c in enumerate(self.centers):
            for xi, x_ in enumerate(x):
                g[xi, ci] = self.basic_func(c, x_)
        return g

    def train(self, x, y):
        """

        :param X:100*3
        :param Y:100*1
        :return
        """
        rnd_idx = np.random.permutation(x.shape[0])[:self.num_centers]
        self.centers = [x[i, :] for i in rnd_idx]
        g = self.act(x)
        self.w = np.dot(pinv(g), y)
        pass

    def predict(self, x):
        g = self.act(x)
        y = np.dot(g, self.w)
        return y

data = xlrd2.open_workbook('123.xls')
table = data.sheet_by_name('Sheet')
row_num = table.nrows
col_num = table.ncols

n = 60
x_tr = np.zeros((n, 3))
x_ts = np.zeros((n, 3))
for i in range(row_num - n):
    for j in range(col_num - 1):
        x_tr[i, j] = table.cell(i, j).value

for i in range(n, row_num):
    for j in range(col_num - 1):
        x_ts[i - n, j] = table.cell(i, j).value

y_tr = np.zeros((n, 1))
y_ts = np.zeros((n, 1))
for i in range(row_num - n):
    y_tr[i, 0] = table.cell(i, 3).value
    pass

for i in range(n, row_num):
    y_ts[i - n, 0] = table.cell(i, 3).value

rbf = RBF(3, 50000, 1)
rbf.train(x_tr, y_tr)
z = rbf.predict(x_ts)

wb = op.Workbook()
ws = wb['Sheet']


for i in range(1, 61):
    ws.cell(row=i, column=1).value = y_tr[i - 1, 0]

for i in range(1, 61):
    ws.cell(row=i, column=2).value = z[i - 1, 0]

wb.save('对比.xls')

def MSE(x, y):
    l = len(x)
    res = 0
    for i in range(l):
        res = res + (x[i] - y[i]) ** 2
        pass
    res = res / l
    return res

def Re(x, y):
    l = len(x)
    res = 0
    for i in range(l):
        if x[i] != 0:
            res = res + abs((x[i] - y[i]) / x[i])
    res = res / l
    return res

result1 = MSE(y_tr, z)
print("预测值与真实值的MSE为：", result1)
result2 = Re(y_tr, z)
print("平均相对误差", result2)

plt.plot(range(1, 61), y_ts, "k-", label='real')
plt.legend()
plt.ylim((0, 0.4))
# plt.savefig(r'真实.jpg')
# plt.show()
plt.plot(range(1, 61), z, "r-", label='predict')
plt.legend()
plt.ylim((0, 0.4))
plt.savefig(r'对比图.jpg')
plt.show()